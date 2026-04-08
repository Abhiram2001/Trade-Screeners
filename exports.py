"""Export helpers: TradingView watchlist, TWS CSV, and Excel workbook."""

from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font as XFont, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Columns that contain numeric percentage or price values
_NUM_COLS = {
    "Price", "Open", "Gap_Supp%", "Gap_Cross%", "Gap_Pull%",
    "Perf_W%", "RSI", "MktCap_B", "Change%",
    "Day_High", "Day_Low", "Wk_High", "Wk_Low",
}

_TV_ORDER = ["NASDAQ", "NYSE", "AMEX", "BATS"]


def export_tradingview(results: list, watchlist_name: str, path: str):
    """Write a TradingView-importable watchlist .txt file.

    Format (all comma-separated, ###SectionName as inline separators):
        ###watchlist name=Potential_Bullish_1
        ###Stocks,NASDAQ:AAPL,NASDAQ:TSLA,NYSE:GS,...
    """
    grouped: dict = {}
    for r in results:
        exc = (r.get("Exchange") or "AMEX").strip().upper()
        sym = r["Symbol"].strip()
        grouped.setdefault(exc, []).append(f"{exc}:{sym}")

    parts = ["###Stocks"]
    for exc in _TV_ORDER:
        parts.extend(grouped.pop(exc, []))
    for exc in sorted(grouped):
        parts.extend(grouped[exc])

    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write(f"###watchlist name={watchlist_name}\n")
        f.write(",".join(parts) + "\n")


def export_tws(results: list, path: str):
    """Write an Interactive Brokers Trader Workstation (TWS) importable watchlist CSV.

    Format:
        SYM,AAPL,SMART/AMEX
        SYM,GS,SMART/AMEX
    Exchange routing is always fixed to SMART/AMEX.
    """
    lines = [f"SYM,{r['Symbol'].strip()},SMART/AMEX" for r in results]
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines) + "\n")


def export_excel(results: list, job_name: str, screener_id: int, job_num: int, path: str):
    """Write a formatted Excel workbook with a data sheet and a TradingView import sheet."""
    wb = openpyxl.Workbook()

    # ── Shared styles ─────────────────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="1C2038")
    hdr_font   = XFont(bold=True, color="9AA3B8", size=10)
    alt_fill   = PatternFill("solid", fgColor="141726")
    reg_fill   = PatternFill("solid", fgColor="0D0F1A")
    pos_font   = XFont(color="34D399", size=11)
    neg_font   = XFont(color="F87171", size=11)
    sym_font   = XFont(bold=True, color="6C8EF7", size=11)
    body_font  = XFont(color="DDE4F0", size=11)
    muted_font = XFont(color="7A86A0", size=10)
    info_font  = XFont(color="7A86A0", italic=True, size=10)
    cell_border = Border(bottom=Side(style="thin", color="1C2038"))

    # ── Sheet 1: Full results ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Screener Results"

    if results:
        cols = list(results[0].keys())

        for c, col in enumerate(cols, 1):
            cell = ws.cell(1, c, col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

        pct_cols = {"Gap_Supp%", "Gap_Cross%", "Gap_Pull%", "Perf_W%", "Change%"}
        for r, row in enumerate(results, 2):
            fill = alt_fill if r % 2 == 0 else reg_fill
            for c, col in enumerate(cols, 1):
                val = row.get(col)
                display = val if val is not None else ""
                cell = ws.cell(r, c, display)
                cell.fill = fill
                cell.border = cell_border
                is_num = isinstance(val, (int, float))
                cell.alignment = Alignment(
                    horizontal="right" if is_num else "left",
                    vertical="center",
                )
                if col == "Symbol":
                    cell.font = sym_font
                elif col in ("Sector", "Exchange"):
                    cell.font = muted_font
                elif col == "MktCap_B" and is_num:
                    cell.font = muted_font
                    cell.number_format = '#,##0.0"B"'
                elif col in pct_cols and is_num:
                    cell.font = pos_font if val > 0 else (neg_font if val < 0 else body_font)
                    cell.number_format = (
                        '+0.00%;-0.00%;0.00%' if col in ("Change%", "Perf_W%") else '0.00%'
                    )
                elif col == "Volume" and is_num:
                    cell.number_format = "#,##0"
                    cell.font = muted_font
                elif col == "RSI" and is_num:
                    cell.font = XFont(
                        color="F87171" if val > 70 else ("FBBF24" if val < 30 else "DDE4F0"),
                        size=11,
                    )
                elif is_num:
                    cell.font = body_font
                    cell.number_format = "#,##0.00"
                else:
                    cell.font = body_font

        for c, col in enumerate(cols, 1):
            max_len = max(len(col), max(
                (len(str(row.get(col, "") or "")) for row in results), default=0
            ))
            ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 32)

        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    ws.sheet_view.showGridLines = False

    # ── Sheet 2: TradingView Import ───────────────────────────────────────────
    ws2 = wb.create_sheet("TradingView Import")
    ws2.sheet_view.showGridLines = False

    hdr2_font = XFont(bold=True, color="6C8EF7", size=11)
    tv_hdr    = XFont(bold=True, color="34D399", size=11)
    tws_hdr   = XFont(bold=True, color="A78BFA", size=11)

    ws2.cell(1, 1, f"Export Formats — {job_name}").font = hdr2_font
    ws2.cell(2, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = info_font

    # Column A: TradingView .txt format
    tv_grouped: dict = {}
    for row in results:
        exc = (row.get("Exchange") or "AMEX").strip().upper()
        sym = row["Symbol"].strip()
        tv_grouped.setdefault(exc, []).append(f"{exc}:{sym}")
    tv_parts = ["###Stocks"]
    for exc in _TV_ORDER:
        tv_parts.extend(tv_grouped.pop(exc, []))
    for exc in sorted(tv_grouped):
        tv_parts.extend(tv_grouped[exc])
    tv_line = ",".join(tv_parts)

    ws2.cell(4, 1, "TradingView Watchlist (.txt)").font = tv_hdr
    ws2.cell(5, 1, "Save column A rows 7-8 as a .txt and import:").font = info_font
    ws2.cell(6, 1, "TradingView → Watchlist → ⋮ → Import watchlist").font = info_font
    ws2.cell(7, 1, f"###watchlist name={job_name}").font = XFont(color="AABCD8", size=10, italic=True)
    ws2.cell(8, 1, tv_line).font = sym_font

    # Column C: TWS .csv format
    ws2.cell(4, 3, "Trader Workstation Watchlist (.csv)").font = tws_hdr
    ws2.cell(5, 3, "Save column C rows as a .csv file and import via:").font = info_font
    ws2.cell(6, 3, "TWS → File → Import Watchlist").font = info_font
    for i, row in enumerate(results, 7):
        ws2.cell(i, 3, f"SYM,{row['Symbol'].strip()},SMART/AMEX").font = sym_font

    # Column E+: Reference table
    ws2.cell(4, 5, "Reference").font = hdr2_font
    for c, hdr in enumerate(["Ticker", "Exchange", "Company", "Price"], 5):
        cell = ws2.cell(5, c, hdr)
        cell.font = hdr2_font
        cell.fill = hdr_fill
    for i, row in enumerate(results, 6):
        ws2.cell(i, 5, row["Symbol"].strip()).font = sym_font
        ws2.cell(i, 6, (row.get("Exchange") or "").strip().upper()).font = muted_font
        ws2.cell(i, 7, row.get("Company", "")).font = body_font
        pc = ws2.cell(i, 8, row.get("Price"))
        pc.font = body_font
        pc.number_format = "#,##0.00"

    for col_letter, width in [("A", 48), ("B", 2), ("C", 28), ("D", 2),
                               ("E", 12), ("F", 12), ("G", 32), ("H", 12)]:
        ws2.column_dimensions[col_letter].width = width

    # ── Sheet 3: Summary ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 30

    title_font = XFont(bold=True, color="DDE4F0", size=14)
    val_font   = XFont(bold=True, color="6C8EF7", size=13)

    ws3.cell(1, 1, f"Trade Screener — {job_name}").font = title_font
    ws3.cell(2, 1, f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}").font = info_font
    ws3.cell(3, 1, f"Screener {screener_id}  ·  Job {job_num}").font = info_font
    ws3.cell(5, 1, "Stocks Matched").font = XFont(color="7A86A0", size=11)
    ws3.cell(5, 2, len(results)).font = val_font

    if results:
        sectors = {}
        for row in results:
            s = row.get("Sector", "Unknown") or "Unknown"
            sectors[s] = sectors.get(s, 0) + 1
        ws3.cell(7, 1, "Sector Breakdown").font = XFont(bold=True, color="9AA3B8", size=11)
        for i, (sec, cnt) in enumerate(sorted(sectors.items(), key=lambda x: -x[1]), 8):
            ws3.cell(i, 1, sec).font = body_font
            ws3.cell(i, 2, cnt).font = val_font

    wb.save(path)
