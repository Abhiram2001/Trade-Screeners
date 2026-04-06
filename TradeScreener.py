#!/usr/bin/env python3
"""
Trade Screener — Desktop Application
=====================================
Unified GUI combining both screeners with fully configurable parameters,
in-app results viewer, and export to Excel + TradingView watchlist.

Screener 1 — Bullish Near Support
  Job 1: Price > EMA(daily trend) + Price > EMA(weekly trend) + EMA(support) ≤ Price ≤ EMA(support)×(1+gap%)
  Job 2: Daily EMA crossover on Job 1 watchlist stocks

Screener 2 — Weekly Bullish / Daily Pullback
  Job 1: Price > EMA(weekly trend) + Price < EMA(daily pullback) + Price > EMA(recovery)
  Job 2: Daily EMA crossover on Job 1 watchlist stocks

Install dependencies:
    pip3 install PyQt5 pandas requests openpyxl tabulate
"""

import json
import os
import ssl
import sys
import urllib.request
import warnings
from dataclasses import dataclass, asdict, field
from datetime import datetime

warnings.filterwarnings("ignore")

# ── Dependency check ───────────────────────────────────────────────────────────
import importlib.util as _iutil
_DEPS = [("pandas", "pandas"), ("requests", "requests"),
         ("PyQt5", "PyQt5"), ("openpyxl", "openpyxl")]
_miss = [n for m, n in _DEPS if _iutil.find_spec(m) is None]
if _miss:
    print(f"Missing packages. Run:\n  pip3 install {' '.join(_miss)}")
    sys.exit(1)

import pandas as pd
import requests
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QFormLayout, QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView,
    QPushButton, QLabel, QLineEdit, QCheckBox, QSpinBox, QDoubleSpinBox,
    QGroupBox, QSplitter, QScrollArea, QFileDialog, QMessageBox,
    QProgressBar, QTextEdit, QAbstractItemView, QSizePolicy, QFrame,
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSortFilterProxyModel
from PyQt5.QtGui import QColor, QFont, QIcon
import openpyxl
from openpyxl.styles import PatternFill, Font as XFont, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants ──────────────────────────────────────────────────────────────────
TV_URL        = "https://scanner.tradingview.com/america/scan"
BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
SETTINGS_FILE = os.path.join(BASE_DIR, "screener_settings.json")
WL1_FILE      = os.path.join(BASE_DIR, "s1_watchlist.json")
WL2_FILE      = os.path.join(BASE_DIR, "s2_watchlist.json")

# ── Stylesheet ─────────────────────────────────────────────────────────────────
STYLE = """
QMainWindow, QWidget {
    background-color: #0d0f1a;
    color: #dde4f0;
    font-size: 13px;
}
QTabWidget::pane {
    border: 1px solid #2a3060;
    background-color: #141726;
}
QTabWidget > QTabBar { alignment: left; }
QTabBar { font-size: 13px; }
QTabBar::tab {
    background-color: #0d0f1a;
    color: #8090a8;
    padding: 10px 30px;
    min-width: 260px;
    border: 1px solid #2a3060;
    border-bottom: none;
    border-radius: 6px 6px 0 0;
    margin-right: 4px;
    font-weight: 600;
    font-size: 13px;
}
QTabBar::tab:selected { background-color: #141726; color: #ffffff; border-color: #6c8ef7; }
QTabBar::tab:hover:!selected { color: #dde4f0; background-color: #141726; }
QLabel { color: #dde4f0; }
QLineEdit, QSpinBox, QDoubleSpinBox {
    background-color: #1a1f3a;
    border: 1px solid #2a3060;
    border-radius: 6px;
    color: #dde4f0;
    padding: 4px 6px;
    min-height: 26px;
    font-size: 12px;
}
QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus {
    border-color: #6c8ef7;
    background-color: #1e2548;
}
QSpinBox::up-button, QSpinBox::down-button,
QDoubleSpinBox::up-button, QDoubleSpinBox::down-button {
    width: 20px;
    background-color: #3a4888;
    border: none;
    border-left: 1px solid #2a3060;
}
QSpinBox::up-button { subcontrol-position: top right; border-radius: 0 4px 0 0; }
QSpinBox::down-button { subcontrol-position: bottom right; border-radius: 0 0 4px 0; }
QDoubleSpinBox::up-button { subcontrol-position: top right; border-radius: 0 4px 0 0; }
QDoubleSpinBox::down-button { subcontrol-position: bottom right; border-radius: 0 0 4px 0; }
QSpinBox::up-button:hover, QSpinBox::down-button:hover,
QDoubleSpinBox::up-button:hover, QDoubleSpinBox::down-button:hover {
    background-color: #5568b8;
}
QSpinBox::up-arrow, QDoubleSpinBox::up-arrow {
    width: 7px;
    height: 5px;
}
QSpinBox::down-arrow, QDoubleSpinBox::down-arrow {
    width: 7px;
    height: 5px;
}
QCheckBox { spacing: 8px; color: #c8d4ec; font-size: 13px; }
QCheckBox::indicator {
    width: 16px; height: 16px;
    border-radius: 4px;
    border: 1px solid #3a4070;
    background-color: #1a1f3a;
}
QCheckBox::indicator:checked { background-color: #6c8ef7; border-color: #6c8ef7; }
QPushButton {
    background-color: #1e2545;
    color: #c8d0e2;
    border: 1px solid #3a4870;
    border-radius: 7px;
    padding: 7px 18px;
    font-weight: 600;
    min-height: 30px;
    font-size: 13px;
}
QPushButton:hover { background-color: #2a3570; border-color: #6c8ef7; color: #ffffff; }
QPushButton:pressed { background-color: #141726; }
QPushButton:disabled { color: #3a4060; border-color: #1c2038; background-color: #0f1220; }
QPushButton#btnPrimary {
    background-color: #3a60e8;
    color: #ffffff;
    border: 2px solid #5a80f8;
    font-weight: 700;
    font-size: 13px;
}
QPushButton#btnPrimary:hover { background-color: #5a78f0; border-color: #8aa5ff; }
QPushButton#btnPrimary:pressed { background-color: #2a50d0; }
QPushButton#btnPrimary:disabled { background-color: #1e2545; color: #3a4060; border-color: #252a45; }
QPushButton#btnDanger { background-color: #5a1818; color: #ff9090; border: 1px solid #8a2828; }
QPushButton#btnDanger:hover { background-color: #8a2828; color: #ffffff; }
QPushButton#btnSuccess { background-color: #163828; color: #34d399; border: 1px solid #1e5038; }
QPushButton#btnSuccess:hover { background-color: #1e5038; color: #5de8ae; }
QPushButton#btnTV { background-color: #162818; color: #34d399; border: 1px solid #1e4028; }
QPushButton#btnTV:hover { background-color: #1e4028; color: #5de8ae; }
QTableWidget {
    background-color: #141726;
    border: none;
    gridline-color: #1c2038;
    color: #dde4f0;
    font-size: 12px;
    selection-background-color: #252a45;
    alternate-background-color: #161929;
}
QTableWidget QHeaderView::section {
    background-color: #1c2038;
    color: #8892aa;
    border: none;
    border-right: 1px solid #252a45;
    border-bottom: 1px solid #252a45;
    padding: 7px 10px;
    font-size: 11px;
    font-weight: 700;
}
QTableWidget QHeaderView::section:hover { color: #dde4f0; background-color: #252a45; }
QTableWidget::item { padding: 5px 8px; border-bottom: 1px solid #1c2038; }
QTableWidget::item:selected { background-color: #252a45; }
QScrollBar:vertical { background-color: #0d0f1a; width: 8px; margin: 0; }
QScrollBar::handle:vertical { background-color: #2a3060; border-radius: 4px; min-height: 30px; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
QScrollBar:horizontal { background-color: #0d0f1a; height: 8px; margin: 0; }
QScrollBar::handle:horizontal { background-color: #2a3060; border-radius: 4px; min-width: 30px; }
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; }
QStatusBar {
    background-color: #080a12;
    color: #7a86a0;
    border-top: 1px solid #1c2038;
    padding: 4px 12px;
    font-size: 12px;
}
QProgressBar {
    background-color: #141726;
    border: 1px solid #252a45;
    border-radius: 4px;
    color: transparent;
    max-height: 6px;
}
QProgressBar::chunk { background-color: #6c8ef7; border-radius: 4px; }
QSplitter::handle { background-color: #252a45; }
QSplitter::handle:horizontal { width: 1px; }
QScrollArea { border: none; background: transparent; }
QScrollArea > QWidget > QWidget { background: transparent; }
QTextEdit {
    background-color: #080a12;
    color: #6a7890;
    border: none;
    border-top: 1px solid #1c2038;
    font-family: 'Menlo', 'Consolas', monospace;
    font-size: 11px;
}
"""

# ── Config dataclasses ─────────────────────────────────────────────────────────

@dataclass
class S1J1Cfg:
    include_sp500    : bool  = True
    include_ndx100   : bool  = True
    ema_daily_trend  : int   = 200    # Price > EMA(n) Daily
    ema_weekly_trend : int   = 200    # Price > EMA(n) Weekly
    ema_support      : int   = 20     # Near-support EMA (gap check)
    gap_max_pct      : float = 3.0    # Max % above support EMA
    perf_weekly_min  : float = 0.0    # Weekly performance floor
    min_mktcap_b     : float = 0.0    # Min market cap ($B), 0 = no filter

@dataclass
class S1J2Cfg:
    ema_crossover : int = 20          # Daily EMA crossover period

@dataclass
class S2J1Cfg:
    include_sp500      : bool  = True
    include_ndx100     : bool  = True
    ema_weekly_trend   : int   = 200  # Price > EMA(n) Weekly
    ema_daily_pullback : int   = 200  # Price < EMA(n) Daily
    ema_recovery       : int   = 8    # Price > EMA(n) Daily
    min_mktcap_b       : float = 0.0

@dataclass
class S2J2Cfg:
    ema_crossover : int = 20

# ── Settings persistence ───────────────────────────────────────────────────────

def load_settings() -> dict:
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_settings(data: dict):
    try:
        with open(SETTINGS_FILE, "w") as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass

# ── TradingView API ────────────────────────────────────────────────────────────

def get_index_tickers() -> tuple:
    """Return (sp500_set, nasdaq100_set)."""
    ctx = ssl._create_unverified_context()
    def _fetch(url, idx, col):
        req  = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        html = urllib.request.urlopen(req, context=ctx).read()
        return set(pd.read_html(html)[idx][col].tolist())
    sp500  = _fetch("https://en.wikipedia.org/wiki/List_of_S%26P_500_companies", 0, "Symbol")
    nasdaq = _fetch("https://en.wikipedia.org/wiki/Nasdaq-100", 4, "Ticker")
    return sp500, nasdaq

def tv_scan(filters: list, columns: list, limit: int = 3000) -> list:
    # Sanitize filter values: ensure numeric "right" values are JSON-safe
    clean_filters = []
    for f in filters:
        cf = dict(f)
        if isinstance(cf.get("right"), float) and cf["right"] == int(cf["right"]):
            cf["right"] = int(cf["right"])
        clean_filters.append(cf)
    payload = {
        "filter":   clean_filters,
        "options":  {"lang": "en"},
        "markets":  ["america"],
        "screener": "america",
        "symbols":  {"query": {"types": []}, "tickers": []},
        "columns":  columns,
        "sort":     {"sortBy": "market_cap_basic", "sortOrder": "desc"},
        "range":    [0, limit],
    }
    r = requests.post(TV_URL, json=payload, timeout=30)
    if not r.ok:
        raise requests.HTTPError(
            f"{r.status_code} {r.reason} — TradingView response: {r.text[:300]}",
            response=r,
        )
    return r.json().get("data", [])

# ── Screener logic ─────────────────────────────────────────────────────────────

def screener1_job1(cfg: S1J1Cfg, universe: set) -> list:
    dt = f"EMA{cfg.ema_daily_trend}"
    wt = f"EMA{cfg.ema_weekly_trend}|1W"
    ds = f"EMA{cfg.ema_support}"
    cols = [
        "name", "description", "close",
        dt, wt, ds,
        "Perf.W", "RSI", "volume", "market_cap_basic",
        "sector", "change", "high", "low", "high|1W", "low|1W", "exchange",
    ]
    filt = [
        {"left": dt,       "operation": "less",    "right": "close"},
        {"left": wt,       "operation": "less",    "right": "close"},
        {"left": ds,       "operation": "less",    "right": "close"},
        {"left": "Perf.W", "operation": "greater", "right": cfg.perf_weekly_min},
    ]
    raw, out = tv_scan(filt, cols), []
    for item in raw:
        d = item["d"]
        try:
            sym, price, ema_s = d[0], (d[2] or 0), d[5]
            if sym not in universe or not price or not ema_s:
                continue
            gap = (price - ema_s) / ema_s * 100
            if gap > cfg.gap_max_pct:
                continue
            mc = d[9]
            if cfg.min_mktcap_b > 0 and (not mc or mc / 1e9 < cfg.min_mktcap_b):
                continue
            out.append({
                "Symbol":                              sym,
                "Exchange":                            d[16] or "",
                "Company":                             (d[1] or "")[:35],
                "Price":                               round(price, 2),
                f"EMA{cfg.ema_support}(1D)":           round(ema_s, 2),
                "Gap_Supp%":                           round(gap, 2),
                f"EMA{cfg.ema_daily_trend}(1D)":       round(d[3], 2) if d[3] else None,
                f"EMA{cfg.ema_weekly_trend}(1W)":      round(d[4], 2) if d[4] else None,
                "Perf_W%":                             round(d[6], 2) if d[6] is not None else None,
                "RSI":                                 round(d[7], 1) if d[7] else None,
                "Volume":                              int(d[8]) if d[8] else None,
                "MktCap_B":                            round(mc / 1e9, 1) if mc else None,
                "Sector":                              d[10] or "",
                "Change%":                             round(d[11], 2) if d[11] is not None else None,
                "Day_High":                            round(d[12], 2) if d[12] else None,
                "Day_Low":                             round(d[13], 2) if d[13] else None,
                "Wk_High":                             round(d[14], 2) if d[14] else None,
                "Wk_Low":                              round(d[15], 2) if d[15] else None,
            })
        except Exception:
            continue
    return out


def screener1_job2(cfg: S1J2Cfg, watchlist: set, universe: set) -> list:
    ec = f"EMA{cfg.ema_crossover}"
    cols = [
        "name", "description", "close", "open", ec,
        "EMA200", "EMA200|1W", "RSI", "volume", "market_cap_basic",
        "sector", "change", "high", "low", "high|1W", "low|1W", "exchange",
    ]
    filt = [{"left": "close", "operation": "crosses", "right": ec}]
    raw, out = tv_scan(filt, cols), []
    for item in raw:
        d = item["d"]
        try:
            sym, price, op, ema_c = d[0], (d[2] or 0), (d[3] or 0), d[4]
            if sym not in watchlist or not ema_c:
                continue
            if not (price > ema_c and op <= ema_c):
                continue
            mc = d[9]
            out.append({
                "Symbol":                          sym,
                "Exchange":                        d[16] or "",
                "Company":                         (d[1] or "")[:35],
                "Price":                           round(price, 2),
                "Open":                            round(op, 2),
                f"EMA{cfg.ema_crossover}(1D)":     round(ema_c, 2),
                "Gap_Cross%":                      round((price - ema_c) / ema_c * 100, 2),
                "EMA200(1D)":                      round(d[5], 2) if d[5] else None,
                "EMA200(1W)":                      round(d[6], 2) if d[6] else None,
                "RSI":                             round(d[7], 1) if d[7] else None,
                "Volume":                          int(d[8]) if d[8] else None,
                "MktCap_B":                        round(mc / 1e9, 1) if mc else None,
                "Sector":                          d[10] or "",
                "Change%":                         round(d[11], 2) if d[11] is not None else None,
                "Day_High":                        round(d[12], 2) if d[12] else None,
                "Day_Low":                         round(d[13], 2) if d[13] else None,
                "Wk_High":                         round(d[14], 2) if d[14] else None,
                "Wk_Low":                          round(d[15], 2) if d[15] else None,
            })
        except Exception:
            continue
    return out


def screener2_job1(cfg: S2J1Cfg, universe: set) -> list:
    wt  = f"EMA{cfg.ema_weekly_trend}|1W"
    dpb = f"EMA{cfg.ema_daily_pullback}"
    rec = f"EMA{cfg.ema_recovery}"
    cols = [
        "name", "description", "close",
        wt, dpb, rec,
        "Perf.W", "RSI", "volume", "market_cap_basic",
        "sector", "change", "high", "low", "high|1W", "low|1W", "exchange",
    ]
    filt = [
        {"left": wt,    "operation": "less", "right": "close"},
        {"left": "close", "operation": "less", "right": dpb},
        {"left": rec,   "operation": "less", "right": "close"},
    ]
    raw, out = tv_scan(filt, cols), []
    for item in raw:
        d = item["d"]
        try:
            sym, price = d[0], (d[2] or 0)
            ema_wt, ema_dpb, ema_rec = d[3], d[4], d[5]
            if sym not in universe or not price:
                continue
            if not ema_wt or not ema_dpb or not ema_rec:
                continue
            mc = d[9]
            if cfg.min_mktcap_b > 0 and (not mc or mc / 1e9 < cfg.min_mktcap_b):
                continue
            out.append({
                "Symbol":                              sym,
                "Exchange":                            d[16] or "",
                "Company":                             (d[1] or "")[:35],
                "Price":                               round(price, 2),
                f"EMA{cfg.ema_recovery}(1D)":          round(ema_rec, 2),
                f"EMA{cfg.ema_daily_pullback}(1D)":    round(ema_dpb, 2),
                "Gap_Pull%":                           round((price - ema_dpb) / ema_dpb * 100, 2),
                f"EMA{cfg.ema_weekly_trend}(1W)":      round(ema_wt, 2),
                "Perf_W%":                             round(d[6], 2) if d[6] is not None else None,
                "RSI":                                 round(d[7], 1) if d[7] else None,
                "Volume":                              int(d[8]) if d[8] else None,
                "MktCap_B":                            round(mc / 1e9, 1) if mc else None,
                "Sector":                              d[10] or "",
                "Change%":                             round(d[11], 2) if d[11] is not None else None,
                "Day_High":                            round(d[12], 2) if d[12] else None,
                "Day_Low":                             round(d[13], 2) if d[13] else None,
                "Wk_High":                             round(d[14], 2) if d[14] else None,
                "Wk_Low":                              round(d[15], 2) if d[15] else None,
            })
        except Exception:
            continue
    return out


def screener2_job2(cfg: S2J2Cfg, watchlist: set, universe: set) -> list:
    ec = f"EMA{cfg.ema_crossover}"
    cols = [
        "name", "description", "close", "open", ec,
        "EMA200", "EMA200|1W", "EMA8", "RSI", "volume", "market_cap_basic",
        "sector", "change", "high", "low", "high|1W", "low|1W", "exchange",
    ]
    filt = [{"left": "close", "operation": "crosses", "right": ec}]
    raw, out = tv_scan(filt, cols), []
    for item in raw:
        d = item["d"]
        try:
            sym, price, op, ema_c = d[0], (d[2] or 0), (d[3] or 0), d[4]
            if sym not in watchlist or not ema_c:
                continue
            if not (price > ema_c and op <= ema_c):
                continue
            mc = d[10]
            out.append({
                "Symbol":                          sym,
                "Exchange":                        d[17] or "",
                "Company":                         (d[1] or "")[:35],
                "Price":                           round(price, 2),
                "Open":                            round(op, 2),
                f"EMA{cfg.ema_crossover}(1D)":     round(ema_c, 2),
                "Gap_Cross%":                      round((price - ema_c) / ema_c * 100, 2),
                "EMA200(1D)":                      round(d[5], 2) if d[5] else None,
                "EMA200(1W)":                      round(d[6], 2) if d[6] else None,
                "EMA8(1D)":                        round(d[7], 2) if d[7] else None,
                "RSI":                             round(d[8], 1) if d[8] else None,
                "Volume":                          int(d[9]) if d[9] else None,
                "MktCap_B":                        round(mc / 1e9, 1) if mc else None,
                "Sector":                          d[11] or "",
                "Change%":                         round(d[12], 2) if d[12] is not None else None,
                "Day_High":                        round(d[13], 2) if d[13] else None,
                "Day_Low":                         round(d[14], 2) if d[14] else None,
                "Wk_High":                         round(d[15], 2) if d[15] else None,
                "Wk_Low":                          round(d[16], 2) if d[16] else None,
            })
        except Exception:
            continue
    return out

# ── Export helpers ─────────────────────────────────────────────────────────────

_NUM_COLS = {
    "Price", "Open", "Gap_Supp%", "Gap_Cross%", "Gap_Pull%",
    "Perf_W%", "RSI", "MktCap_B", "Change%",
    "Day_High", "Day_Low", "Wk_High", "Wk_Low",
}

def export_tradingview(results: list, watchlist_name: str, path: str):
    """Write a TradingView-importable watchlist CSV file.

    Format matches the TradingView watchlist export format:
        COLUMN,0,,
        DES,AAPL,STK,SMART/AMEX
        DES,MSFT,STK,SMART/AMEX
        ,,,
    """
    # Exchange name → TradingView routing suffix mapping
    _EXC_MAP = {
        "NASDAQ": "SMART/NASDAQ",
        "NYSE":   "SMART/NYSE",
        "AMEX":   "SMART/AMEX",
        "BATS":   "SMART/BATS",
    }

    lines = [f"COLUMN,0,,"]
    for r in results:
        exc_raw = (r.get("Exchange") or "").strip().upper()
        routing = _EXC_MAP.get(exc_raw, "SMART/AMEX")
        sym = r["Symbol"].strip()
        lines.append(f"DES,{sym},STK,{routing}")
    lines.append(",,,")

    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines) + "\n")


def export_excel(results: list, job_name: str, screener_id: int, job_num: int, path: str):
    """Write a formatted Excel workbook with a data sheet and a TradingView import sheet."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Full results ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Screener Results"

    hdr_fill   = PatternFill("solid", fgColor="1C2038")
    hdr_font   = XFont(bold=True, color="9AA3B8", size=10)
    alt_fill   = PatternFill("solid", fgColor="141726")
    reg_fill   = PatternFill("solid", fgColor="0D0F1A")
    pos_font   = XFont(color="34D399", size=11)
    neg_font   = XFont(color="F87171", size=11)
    sym_font   = XFont(bold=True, color="6C8EF7", size=11)
    body_font  = XFont(color="DDE4F0", size=11)
    muted_font = XFont(color="7A86A0", size=10)

    thin = Side(style="thin", color="252A45")
    cell_border = Border(bottom=Side(style="thin", color="1C2038"))

    if results:
        cols = list(results[0].keys())

        # Header row
        for c, col in enumerate(cols, 1):
            cell = ws.cell(1, c, col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

        # Data rows
        pct_cols = {"Gap_Supp%", "Gap_Cross%", "Gap_Pull%", "Perf_W%", "Change%"}
        for r, row in enumerate(results, 2):
            fill = alt_fill if r % 2 == 0 else reg_fill
            for c, col in enumerate(cols, 1):
                val = row.get(col)
                display = val if val is not None else ""
                cell = ws.cell(r, c, display)
                cell.fill = fill
                cell.border = cell_border

                is_num = isinstance(val, (int, float))
                cell.alignment = Alignment(
                    horizontal="right" if is_num else "left",
                    vertical="center",
                )

                if col == "Symbol":
                    cell.font = sym_font
                elif col in ("Sector", "Exchange"):
                    cell.font = muted_font
                elif col == "MktCap_B" and is_num:
                    cell.font = muted_font
                    cell.number_format = '#,##0.0"B"'
                elif col in pct_cols and is_num:
                    cell.font = pos_font if val > 0 else (neg_font if val < 0 else body_font)
                    cell.number_format = '+0.00%;-0.00%;0.00%' if col in ("Change%","Perf_W%") else '0.00%'
                elif col == "Volume" and is_num:
                    cell.number_format = "#,##0"
                    cell.font = muted_font
                elif col == "RSI" and is_num:
                    cell.font = XFont(color="F87171" if val > 70 else ("FBBF24" if val < 30 else "DDE4F0"), size=11)
                elif is_num:
                    cell.font = body_font
                    cell.number_format = "#,##0.00"
                else:
                    cell.font = body_font

        # Column widths
        for c, col in enumerate(cols, 1):
            max_len = max(len(col), max(
                (len(str(row.get(col, "") or "")) for row in results), default=0
            ))
            ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 32)

        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    ws.sheet_view.showGridLines = False

    # ── Sheet 2: TradingView Import ───────────────────────────────────────────
    ws2 = wb.create_sheet("TradingView Import")
    ws2.sheet_view.showGridLines = False

    _EXC_MAP = {"NASDAQ": "SMART/NASDAQ", "NYSE": "SMART/NYSE",
                "AMEX": "SMART/AMEX", "BATS": "SMART/BATS"}

    info_font = XFont(color="7A86A0", italic=True, size=10)
    hdr2_font = XFont(bold=True, color="6C8EF7", size=11)

    ws2.cell(1, 1, f"TradingView Watchlist — {job_name}").font = hdr2_font
    ws2.cell(2, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = info_font
    ws2.cell(3, 1, "Import: TradingView → Watchlist → ⋮ → Import watchlist → select the .csv file").font = info_font
    ws2.cell(4, 1, "").font = info_font

    # Column A contains the exact CSV content to save as .csv and import
    ws2.cell(5, 1, "── CSV content (save column A as .csv to import) ──").font = info_font
    ws2.cell(6, 1, "COLUMN,0,,").font = XFont(color="AABCD8", size=11)

    csv_row = 7
    for r, row in enumerate(results, 1):
        exc_raw = (row.get("Exchange") or "").strip().upper()
        routing = _EXC_MAP.get(exc_raw, "SMART/AMEX")
        sym = row["Symbol"].strip()
        csv_line = f"DES,{sym},STK,{routing}"
        ws2.cell(csv_row, 1, csv_line).font = sym_font

        # Reference columns
        ws2.cell(csv_row, 2, sym).font = body_font
        ws2.cell(csv_row, 3, exc_raw or "AMEX").font = muted_font
        ws2.cell(csv_row, 4, row.get("Company", "")).font = body_font
        price_cell = ws2.cell(csv_row, 5, row.get("Price"))
        price_cell.font = body_font
        price_cell.number_format = "#,##0.00"
        csv_row += 1

    ws2.cell(csv_row, 1, ",,,").font = info_font

    for c, hdr in enumerate(["CSV Line (import-ready)", "Ticker", "Exchange", "Company", "Price"], 1):
        cell = ws2.cell(5, c, hdr)
        cell.font = hdr2_font
        cell.fill = hdr_fill

    for col_letter, width in [("A", 32), ("B", 12), ("C", 14), ("D", 35), ("E", 12)]:
        ws2.column_dimensions[col_letter].width = width

    # ── Sheet 3: Summary ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 30

    title_font = XFont(bold=True, color="DDE4F0", size=14)
    val_font   = XFont(bold=True, color="6C8EF7", size=13)

    ws3.cell(1, 1, f"Trade Screener — {job_name}").font = title_font
    ws3.cell(2, 1, f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}").font = info_font
    ws3.cell(3, 1, f"Screener {screener_id}  ·  Job {job_num}").font = info_font
    ws3.cell(5, 1, "Stocks Matched").font = XFont(color="7A86A0", size=11)
    ws3.cell(5, 2, len(results)).font = val_font

    if results:
        sectors = {}
        for row in results:
            s = row.get("Sector", "Unknown") or "Unknown"
            sectors[s] = sectors.get(s, 0) + 1
        ws3.cell(7, 1, "Sector Breakdown").font = XFont(bold=True, color="9AA3B8", size=11)
        for i, (sec, cnt) in enumerate(sorted(sectors.items(), key=lambda x: -x[1]), 8):
            ws3.cell(i, 1, sec).font = body_font
            ws3.cell(i, 2, cnt).font = val_font

    wb.save(path)


# ── Worker thread ──────────────────────────────────────────────────────────────

class ScreenerWorker(QThread):
    log     = pyqtSignal(str)
    j1_done = pyqtSignal(list)
    j2_done = pyqtSignal(list)
    done    = pyqtSignal()
    error   = pyqtSignal(str)

    def __init__(self, run_j1, run_j2, j1_func, j2_func,
                 j1_cfg, j2_cfg, wl_file, screener_id):
        super().__init__()
        self.run_j1 = run_j1
        self.run_j2 = run_j2
        self.j1_func = j1_func
        self.j2_func = j2_func
        self.j1_cfg = j1_cfg
        self.j2_cfg = j2_cfg
        self.wl_file = wl_file
        self.screener_id = screener_id

    def run(self):
        # Build universe
        try:
            self.log.emit("Fetching index constituents from Wikipedia…")
            sp500, nasdaq = get_index_tickers()
            self.log.emit(f"S&P 500: {len(sp500)}  |  Nasdaq 100: {len(nasdaq)}")
        except Exception as e:
            self.error.emit(f"Failed to load universe: {e}")
            return

        j1_cfg = self.j1_cfg
        universe = set()
        if getattr(j1_cfg, "include_sp500", True):
            universe |= sp500
        if getattr(j1_cfg, "include_ndx100", True):
            universe |= nasdaq
        self.log.emit(f"Universe: {len(universe)} unique tickers")

        if self.run_j1:
            try:
                self.log.emit("Running Job 1 — fetching from TradingView…")
                results = self.j1_func(j1_cfg, universe)
                with open(self.wl_file, "w") as f:
                    json.dump([r["Symbol"] for r in results], f, indent=2)
                self.log.emit(f"Job 1 complete: {len(results)} stocks matched ✓")
                self.j1_done.emit(results)
            except requests.HTTPError as e:
                self.error.emit(f"Job 1 API error: {e}")
                return
            except Exception as e:
                self.error.emit(f"Job 1 failed: {e}")
                return

        if self.run_j2:
            try:
                if not os.path.exists(self.wl_file):
                    self.error.emit("Job 1 watchlist not found — run Job 1 first.")
                    self.done.emit()
                    return
                with open(self.wl_file) as f:
                    watchlist = set(json.load(f))
                if not watchlist:
                    self.log.emit("Job 1 watchlist is empty — skipping Job 2.")
                    self.done.emit()
                    return
                self.log.emit(f"Running Job 2 — {len(watchlist)} symbols from Job 1…")
                results = self.j2_func(self.j2_cfg, watchlist, universe)
                self.log.emit(f"Job 2 complete: {len(results)} stocks matched ✓")
                self.j2_done.emit(results)
            except Exception as e:
                self.error.emit(f"Job 2 failed: {e}")

        self.done.emit()


# ── Results table widget ───────────────────────────────────────────────────────

class ResultsTable(QTableWidget):

    # Columns that should be right-aligned / numeric
    _RIGHT_ALIGN = {
        "Price", "Open", "Gap_Supp%", "Gap_Cross%", "Gap_Pull%",
        "Perf_W%", "RSI", "MktCap_B", "Change%",
        "Day_High", "Day_Low", "Wk_High", "Wk_Low", "Volume",
    }
    _PCT_POS   = {"Gap_Supp%", "Gap_Cross%"}   # always green
    _PCT_SIGN  = {"Change%", "Perf_W%"}         # green/red by sign
    _PCT_NEG   = {"Gap_Pull%"}                  # negative = expected (red)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAlternatingRowColors(True)
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setSortingEnabled(True)
        self.horizontalHeader().setStretchLastSection(True)
        self.horizontalHeader().setSortIndicatorShown(True)
        self.verticalHeader().setVisible(False)
        self.setShowGrid(False)
        self._all_data: list = []
        self._cols: list = []

    def set_data(self, rows: list):
        self._all_data = rows
        if not rows:
            self.setRowCount(0)
            self.setColumnCount(0)
            self._cols = []
            return
        self._cols = list(rows[0].keys())
        self._populate(rows)

    def _populate(self, rows: list):
        self.setSortingEnabled(False)
        self.clearContents()
        self.setColumnCount(len(self._cols))
        self.setHorizontalHeaderLabels(self._cols)
        self.setRowCount(len(rows))

        for r, row in enumerate(rows):
            self.setRowHeight(r, 28)
            for c, col in enumerate(self._cols):
                val = row.get(col)
                item = QTableWidgetItem()

                # Store raw numeric value for proper sorting
                if isinstance(val, (int, float)) and val is not None:
                    item.setData(Qt.DisplayRole, val)
                else:
                    item.setText("—" if val is None else str(val))

                # Alignment
                if col in self._RIGHT_ALIGN:
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                else:
                    item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)

                # Color coding
                if col == "Symbol":
                    item.setForeground(QColor("#6c8ef7"))
                    f = QFont()
                    f.setBold(True)
                    item.setFont(f)
                elif col in ("Sector", "Exchange"):
                    item.setForeground(QColor("#7a86a0"))
                elif col == "MktCap_B":
                    item.setForeground(QColor("#7a86a0"))
                elif col in self._PCT_POS and isinstance(val, float):
                    item.setForeground(QColor("#34d399"))
                elif col in self._PCT_SIGN and isinstance(val, float):
                    item.setForeground(QColor("#34d399") if val > 0 else QColor("#f87171"))
                elif col in self._PCT_NEG and isinstance(val, float):
                    # Negative gap = below EMA (pullback — expected, show in amber)
                    item.setForeground(QColor("#fbbf24") if val < 0 else QColor("#34d399"))
                elif col == "RSI" and isinstance(val, float):
                    if val > 70:
                        item.setForeground(QColor("#f87171"))
                    elif val < 30:
                        item.setForeground(QColor("#fbbf24"))

                self.setItem(r, c, item)

        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.horizontalHeader().setStretchLastSection(True)
        self.setSortingEnabled(True)

    def filter_rows(self, text: str) -> int:
        text = text.lower()
        visible = 0
        for r in range(self.rowCount()):
            match = any(
                text in (self.item(r, c).text().lower() if self.item(r, c) else "")
                for c in range(self.columnCount())
            )
            self.setRowHidden(r, not match)
            if match:
                visible += 1
        return visible

    def visible_count(self) -> int:
        return sum(1 for r in range(self.rowCount()) if not self.isRowHidden(r))


# ── Config panel ───────────────────────────────────────────────────────────────

class ConfigPanel(QScrollArea):
    """Scrollable left sidebar containing all tunable parameters."""

    def __init__(self, screener_id: int, saved: dict, parent=None):
        super().__init__(parent)
        self.screener_id = screener_id
        self.setWidgetResizable(True)
        self.setMinimumWidth(300)
        self.setMaximumWidth(360)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        container = QWidget()
        container.setStyleSheet("background-color: #0d0f1a;")
        lay = QVBoxLayout(container)
        lay.setSpacing(12)
        lay.setContentsMargins(10, 12, 10, 20)

        # ── Universe section ──────────────────────────────────────────────────
        lay.addWidget(self._section_hdr("Universe"))
        univ_box = self._section_box()
        ul = QVBoxLayout(univ_box)
        ul.setSpacing(8)
        ul.setContentsMargins(10, 10, 10, 10)
        self.cb_sp500  = QCheckBox("S&P 500  (≈503 stocks)")
        self.cb_ndx100 = QCheckBox("Nasdaq 100  (≈101 stocks)")
        self.cb_sp500.setChecked(saved.get("include_sp500", True))
        self.cb_ndx100.setChecked(saved.get("include_ndx100", True))
        ul.addWidget(self.cb_sp500)
        ul.addWidget(self.cb_ndx100)
        lay.addWidget(univ_box)

        # ── Job 1 section ─────────────────────────────────────────────────────
        if screener_id == 1:
            lay.addWidget(self._section_hdr("Job 1 — Bullish Near Support"))
            j1_box = self._section_box()
            fl = QFormLayout(j1_box)
            fl.setSpacing(9)
            fl.setContentsMargins(10, 10, 10, 10)
            fl.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
            fl.setLabelAlignment(Qt.AlignLeft | Qt.AlignVCenter)

            self.ema_daily_trend  = self._spin(saved.get("ema_daily_trend",  200), 5, 500, "Daily trend EMA period (Price > EMAn Daily)")
            self.ema_weekly_trend = self._spin(saved.get("ema_weekly_trend", 200), 5, 500, "Weekly trend EMA period (Price > EMAn Weekly)")
            self.ema_support      = self._spin(saved.get("ema_support",       20), 3, 200, "Support EMA period (near-support check)")
            self.gap_max_pct      = self._dspin(saved.get("gap_max_pct",      3.0), 0.1, 50.0, 1, "Max % above support EMA (0.1 – 50)")
            self.perf_weekly_min  = self._dspin(saved.get("perf_weekly_min",  0.0), -50.0, 50.0, 1, "Weekly performance floor (%)")
            self.min_mktcap_b     = self._dspin(saved.get("min_mktcap_b",    0.0),  0.0, 5000.0, 1, "Minimum market cap in $B (0 = no filter)")

            fl.addRow(self._lbl("Daily Trend EMA"),  self.ema_daily_trend)
            fl.addRow(self._lbl("Weekly Trend EMA"), self.ema_weekly_trend)
            fl.addRow(self._lbl("Support EMA"),      self.ema_support)
            fl.addRow(self._lbl("Gap Max %"),        self.gap_max_pct)
            fl.addRow(self._lbl("Perf W ≥ %"),       self.perf_weekly_min)
            fl.addRow(self._lbl("Min MktCap $B"),    self.min_mktcap_b)
            lay.addWidget(j1_box)

        else:
            lay.addWidget(self._section_hdr("Job 1 — Wkly Bullish / Pullback"))
            j1_box = self._section_box()
            fl = QFormLayout(j1_box)
            fl.setSpacing(9)
            fl.setContentsMargins(10, 10, 10, 10)
            fl.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
            fl.setLabelAlignment(Qt.AlignLeft | Qt.AlignVCenter)

            self.ema_weekly_trend   = self._spin(saved.get("ema_weekly_trend",   200), 5, 500, "Weekly trend EMA (Price must be ABOVE this)")
            self.ema_daily_pullback = self._spin(saved.get("ema_daily_pullback", 200), 5, 500, "Daily pullback EMA (Price must be BELOW this)")
            self.ema_recovery       = self._spin(saved.get("ema_recovery",         8), 3,  50, "Recovery EMA (Price must be ABOVE this)")
            self.min_mktcap_b       = self._dspin(saved.get("min_mktcap_b",      0.0), 0.0, 5000.0, 1, "Minimum market cap in $B (0 = no filter)")

            fl.addRow(self._lbl("Weekly Trend EMA"),   self.ema_weekly_trend)
            fl.addRow(self._lbl("Daily Pullback EMA"), self.ema_daily_pullback)
            fl.addRow(self._lbl("Recovery EMA"),       self.ema_recovery)
            fl.addRow(self._lbl("Min MktCap $B"),      self.min_mktcap_b)
            lay.addWidget(j1_box)

        # ── Job 2 section ─────────────────────────────────────────────────────
        lay.addWidget(self._section_hdr("Job 2 — EMA Crossover"))
        j2_box = self._section_box()
        fl2 = QFormLayout(j2_box)
        fl2.setSpacing(9)
        fl2.setContentsMargins(10, 10, 10, 10)
        fl2.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        fl2.setLabelAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.ema_crossover = self._spin(saved.get("ema_crossover", 20), 3, 200, "Crossover EMA period (upward daily cross)")
        fl2.addRow(self._lbl("Crossover EMA"), self.ema_crossover)
        lay.addWidget(j2_box)

        # ── Conditions summary ────────────────────────────────────────────────
        lay.addWidget(self._section_hdr("Active Conditions"))
        self.cond_lbl = QLabel()
        self.cond_lbl.setWordWrap(True)
        self.cond_lbl.setStyleSheet(
            "color: #8090b0; font-size: 11px; padding: 10px 8px;"
            "background-color: #111428; border: 1px solid #2e3870; border-radius: 6px;"
        )
        lay.addWidget(self.cond_lbl)
        self._update_conditions()

        # Connect all widgets to update conditions text
        for w in self.findChildren((QSpinBox, QDoubleSpinBox, QCheckBox)):
            if isinstance(w, QCheckBox):
                w.toggled.connect(self._update_conditions)
            else:
                w.valueChanged.connect(self._update_conditions)

        lay.addStretch()
        self.setWidget(container)

    # ── helpers ───────────────────────────────────────────────────────────────

    @staticmethod
    def _section_hdr(text: str) -> QLabel:
        """Visible section title rendered as a plain QLabel — works reliably on macOS."""
        lbl = QLabel(text)
        lbl.setStyleSheet(
            "color: #aabcd8;"
            "background-color: transparent;"
            "font-size: 11px;"
            "font-weight: 700;"
            "padding: 4px 2px 2px 2px;"
        )
        return lbl

    @staticmethod
    def _section_box() -> QFrame:
        """A dark rounded frame used as a section container instead of QGroupBox."""
        f = QFrame()
        f.setStyleSheet(
            "QFrame {"
            "  background-color: #111428;"
            "  border: 1px solid #2e3870;"
            "  border-radius: 7px;"
            "}"
        )
        return f

    @staticmethod
    def _spin(val, mn, mx, tip=""):
        w = QSpinBox()
        w.setRange(mn, mx)
        w.setValue(val)
        w.setToolTip(tip)
        return w

    @staticmethod
    def _dspin(val, mn, mx, dec=2, tip=""):
        w = QDoubleSpinBox()
        w.setRange(mn, mx)
        w.setValue(val)
        w.setDecimals(dec)
        w.setToolTip(tip)
        return w

    @staticmethod
    def _lbl(text):
        l = QLabel(text)
        l.setStyleSheet(
            "color: #c8d4ec;"
            "background-color: transparent;"
            "font-size: 12px;"
            "font-weight: 500;"
        )
        return l

    def _update_conditions(self):
        lines = []
        if self.screener_id == 1:
            dt = getattr(self, "ema_daily_trend",  None)
            wt = getattr(self, "ema_weekly_trend", None)
            ds = getattr(self, "ema_support",      None)
            gp = getattr(self, "gap_max_pct",      None)
            pw = getattr(self, "perf_weekly_min",  None)
            mc = getattr(self, "min_mktcap_b",     None)
            ec = self.ema_crossover
            if dt: lines.append(f"J1: Price > EMA{dt.value()} (1D)")
            if wt: lines.append(f"J1: Price > EMA{wt.value()} (1W)")
            if ds and gp:
                lines.append(f"J1: EMA{ds.value()} ≤ Price ≤ EMA{ds.value()}×{1 + gp.value()/100:.3f}")
            if pw: lines.append(f"J1: Perf.W ≥ {pw.value():.1f}%")
            if mc and mc.value() > 0: lines.append(f"J1: MktCap ≥ ${mc.value():.1f}B")
            lines.append(f"J2: Daily close crosses EMA{ec.value()} upward")
        else:
            wt  = getattr(self, "ema_weekly_trend",   None)
            dpb = getattr(self, "ema_daily_pullback",  None)
            rec = getattr(self, "ema_recovery",        None)
            mc  = getattr(self, "min_mktcap_b",        None)
            ec  = self.ema_crossover
            if wt:  lines.append(f"J1: Price > EMA{wt.value()} (1W)")
            if dpb: lines.append(f"J1: Price < EMA{dpb.value()} (1D)")
            if rec: lines.append(f"J1: Price > EMA{rec.value()} (1D)")
            if mc and mc.value() > 0: lines.append(f"J1: MktCap ≥ ${mc.value():.1f}B")
            lines.append(f"J2: Daily close crosses EMA{ec.value()} upward")
        self.cond_lbl.setText("Active conditions:\n" + "\n".join(f"  ✓ {l}" for l in lines))

    # ── Config extraction ──────────────────────────────────────────────────────

    def get_s1j1(self) -> S1J1Cfg:
        return S1J1Cfg(
            include_sp500    = self.cb_sp500.isChecked(),
            include_ndx100   = self.cb_ndx100.isChecked(),
            ema_daily_trend  = self.ema_daily_trend.value(),
            ema_weekly_trend = self.ema_weekly_trend.value(),
            ema_support      = self.ema_support.value(),
            gap_max_pct      = self.gap_max_pct.value(),
            perf_weekly_min  = self.perf_weekly_min.value(),
            min_mktcap_b     = self.min_mktcap_b.value(),
        )

    def get_s1j2(self) -> S1J2Cfg:
        return S1J2Cfg(ema_crossover=self.ema_crossover.value())

    def get_s2j1(self) -> S2J1Cfg:
        return S2J1Cfg(
            include_sp500      = self.cb_sp500.isChecked(),
            include_ndx100     = self.cb_ndx100.isChecked(),
            ema_weekly_trend   = self.ema_weekly_trend.value(),
            ema_daily_pullback = self.ema_daily_pullback.value(),
            ema_recovery       = self.ema_recovery.value(),
            min_mktcap_b       = self.min_mktcap_b.value(),
        )

    def get_s2j2(self) -> S2J2Cfg:
        return S2J2Cfg(ema_crossover=self.ema_crossover.value())

    def to_dict(self) -> dict:
        d = {
            "include_sp500":  self.cb_sp500.isChecked(),
            "include_ndx100": self.cb_ndx100.isChecked(),
            "ema_crossover":  self.ema_crossover.value(),
        }
        for attr in ("ema_daily_trend", "ema_weekly_trend", "ema_support",
                     "gap_max_pct", "perf_weekly_min", "min_mktcap_b",
                     "ema_daily_pullback", "ema_recovery"):
            w = getattr(self, attr, None)
            if w is not None:
                d[attr] = w.value()
        return d


# ── Screener tab ───────────────────────────────────────────────────────────────

class ScreenerTab(QWidget):
    """Full tab for one screener: toolbar + config panel + results tabs + log."""

    status_msg = pyqtSignal(str)

    def __init__(self, screener_id: int, parent=None):
        super().__init__(parent)
        self.screener_id = screener_id
        self.wl_file     = WL1_FILE if screener_id == 1 else WL2_FILE
        self.worker      = None
        self.j1_results: list = []
        self.j2_results: list = []

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Toolbar ───────────────────────────────────────────────────────────
        root.addWidget(self._build_toolbar())

        # ── Main splitter: config | results ───────────────────────────────────
        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)

        saved = load_settings().get(f"s{screener_id}", {})
        self.cfg_panel = ConfigPanel(screener_id, saved)
        splitter.addWidget(self.cfg_panel)

        right = QWidget()
        rv = QVBoxLayout(right)
        rv.setContentsMargins(10, 10, 10, 6)
        rv.setSpacing(8)

        # Search + count bar
        bar = QHBoxLayout()
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("🔍  Search symbol, company, sector…")
        self.search_box.textChanged.connect(self._on_filter)
        self.count_lbl = QLabel("—")
        self.count_lbl.setStyleSheet("color:#7a86a0; font-size:12px; min-width:80px;")
        bar.addWidget(self.search_box)
        bar.addWidget(self.count_lbl)
        rv.addLayout(bar)

        # Result tabs
        self.res_tabs = QTabWidget()
        j1_name = "Job 1 — Potential Bullish 1" if screener_id == 1 else "Job 1 — Weekly Bullish / Pullback"
        j2_name = "Job 2 — EMA Crossover Breakout"
        self.t_j1 = ResultsTable()
        self.t_j2 = ResultsTable()
        self.res_tabs.addTab(self.t_j1, j1_name)
        self.res_tabs.addTab(self.t_j2, j2_name)
        self.res_tabs.currentChanged.connect(self._on_tab_change)
        rv.addWidget(self.res_tabs)

        # Export bar
        ebar = QHBoxLayout()
        ebar.addStretch()

        self.btn_save_cfg = QPushButton("💾  Save Settings")
        self.btn_save_cfg.setToolTip("Save current configuration to disk")
        self.btn_save_cfg.clicked.connect(self._save_config)

        self.btn_xl  = QPushButton("📊  Export Excel")
        self.btn_xl.setObjectName("btnSuccess")
        self.btn_xl.setToolTip("Export full results + TradingView import sheet to .xlsx")
        self.btn_xl.clicked.connect(self._export_excel)

        self.btn_tv  = QPushButton("📋  TradingView Watchlist")
        self.btn_tv.setObjectName("btnTV")
        self.btn_tv.setToolTip("Export symbols as TradingView importable .txt file")
        self.btn_tv.clicked.connect(self._export_tv)

        for b in (self.btn_save_cfg, self.btn_xl, self.btn_tv):
            ebar.addWidget(b)
        rv.addLayout(ebar)

        splitter.addWidget(right)
        splitter.setSizes([330, 900])
        root.addWidget(splitter)

        # ── Log panel ─────────────────────────────────────────────────────────
        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setMaximumHeight(90)
        self.log_box.setPlaceholderText("Run log will appear here…")
        root.addWidget(self.log_box)

    # ── Toolbar ────────────────────────────────────────────────────────────────

    def _build_toolbar(self) -> QWidget:
        bar = QFrame()
        bar.setStyleSheet("background:#141726; border-bottom:1px solid #252a45;")
        bar.setFixedHeight(58)
        lay = QHBoxLayout(bar)
        lay.setContentsMargins(16, 0, 16, 0)
        lay.setSpacing(10)

        self.btn_j1   = QPushButton("▶  Run Job 1")
        self.btn_j2   = QPushButton("▶  Run Job 2")
        self.btn_both = QPushButton("▶▶  Run Both")
        self.btn_both.setObjectName("btnPrimary")
        self.btn_stop = QPushButton("■  Stop")
        self.btn_stop.setObjectName("btnDanger")
        self.btn_stop.setEnabled(False)

        self.btn_j1.setToolTip("Run Job 1 only (builds watchlist)")
        self.btn_j2.setToolTip("Run Job 2 only (requires prior Job 1 watchlist)")
        self.btn_both.setToolTip("Run Job 1 then Job 2 sequentially")

        self.btn_j1.clicked.connect(lambda: self._run("job1"))
        self.btn_j2.clicked.connect(lambda: self._run("job2"))
        self.btn_both.clicked.connect(lambda: self._run("both"))
        self.btn_stop.clicked.connect(self._stop)

        self.progress = QProgressBar()
        self.progress.setRange(0, 0)
        self.progress.setVisible(False)
        self.progress.setMaximumWidth(160)
        self.progress.setMaximumHeight(6)

        s_name = "Screener 1" if self.screener_id == 1 else "Screener 2"
        title = QLabel(s_name)
        title.setStyleSheet("color:#7a86a0; font-size:11px; font-weight:600;")

        lay.addWidget(title)
        lay.addSpacing(8)
        for b in (self.btn_j1, self.btn_j2, self.btn_both, self.btn_stop):
            lay.addWidget(b)
        lay.addStretch()
        lay.addWidget(self.progress)
        return bar

    # ── Run / stop ─────────────────────────────────────────────────────────────

    def _run(self, mode: str):
        if self.worker and self.worker.isRunning():
            return

        if self.screener_id == 1:
            j1_cfg  = self.cfg_panel.get_s1j1()
            j2_cfg  = self.cfg_panel.get_s1j2()
            j1_func = screener1_job1
            j2_func = screener1_job2
        else:
            j1_cfg  = self.cfg_panel.get_s2j1()
            j2_cfg  = self.cfg_panel.get_s2j2()
            j1_func = screener2_job1
            j2_func = screener2_job2

        self.worker = ScreenerWorker(
            run_j1      = mode in ("job1", "both"),
            run_j2      = mode in ("job2", "both"),
            j1_func     = j1_func,
            j2_func     = j2_func,
            j1_cfg      = j1_cfg,
            j2_cfg      = j2_cfg,
            wl_file     = self.wl_file,
            screener_id = self.screener_id,
        )
        self.worker.log.connect(self._log)
        self.worker.j1_done.connect(self._on_j1)
        self.worker.j2_done.connect(self._on_j2)
        self.worker.done.connect(self._on_done)
        self.worker.error.connect(self._on_error)

        self._set_busy(True)
        self._log(f"Starting {mode.upper()} …")
        self.worker.start()

    def _stop(self):
        if self.worker and self.worker.isRunning():
            self.worker.terminate()
            self._log("Run cancelled.")
            self._set_busy(False)

    def _set_busy(self, busy: bool):
        for b in (self.btn_j1, self.btn_j2, self.btn_both):
            b.setEnabled(not busy)
        self.btn_stop.setEnabled(busy)
        self.progress.setVisible(busy)

    # ── Slots ──────────────────────────────────────────────────────────────────

    def _log(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_box.append(f'<span style="color:#4a5070">[{ts}]</span> {msg}')
        self.status_msg.emit(msg)

    def _on_j1(self, results: list):
        self.j1_results = results
        self.t_j1.set_data(results)
        self.res_tabs.setCurrentIndex(0)
        self._refresh_count()

    def _on_j2(self, results: list):
        self.j2_results = results
        self.t_j2.set_data(results)
        self.res_tabs.setCurrentIndex(1)
        self._refresh_count()

    def _on_done(self):
        self._set_busy(False)
        self._log("Done.")

    def _on_error(self, msg: str):
        self._log(f'<span style="color:#f87171">ERROR: {msg}</span>')
        self._set_busy(False)

    def _on_filter(self, text: str):
        tbl = self._active_table()
        n   = tbl.filter_rows(text)
        self.count_lbl.setText(f"{n} stocks")

    def _on_tab_change(self):
        self.search_box.clear()
        self._refresh_count()

    def _refresh_count(self):
        n = self._active_table().visible_count()
        self.count_lbl.setText(f"{n} stocks")

    def _active_table(self) -> ResultsTable:
        return self.t_j1 if self.res_tabs.currentIndex() == 0 else self.t_j2

    def _current_job_info(self) -> tuple:
        """Returns (results, job_name, job_num)."""
        idx = self.res_tabs.currentIndex()
        if self.screener_id == 1:
            names = ("Potential_Bullish_1", "Bullish_1_20ema")
        else:
            names = ("Potential_Bullish2", "Bullish2_20ema")
        results  = self.j1_results if idx == 0 else self.j2_results
        job_name = names[idx]
        job_num  = idx + 1
        return results, job_name, job_num

    # ── Config persistence ─────────────────────────────────────────────────────

    def _save_config(self):
        all_s = load_settings()
        all_s[f"s{self.screener_id}"] = self.cfg_panel.to_dict()
        save_settings(all_s)
        self._log("Settings saved.")
        self.status_msg.emit("Settings saved ✓")

    # ── Exports ────────────────────────────────────────────────────────────────

    def _export_excel(self):
        results, job_name, job_num = self._current_job_info()
        if not results:
            QMessageBox.information(self, "No Data",
                "No results to export.\nRun the screener first.")
            return
        date_str = datetime.now().strftime("%Y-%m-%d")
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel Report",
            os.path.join(BASE_DIR, f"{job_name}_{date_str}.xlsx"),
            "Excel Workbook (*.xlsx)",
        )
        if not path:
            return
        try:
            export_excel(results, job_name, self.screener_id, job_num, path)
            self._log(f"Excel saved → {path}")
            QMessageBox.information(self, "Export Successful",
                f"Excel workbook saved:\n{path}\n\n"
                "Sheets included:\n"
                "  • Screener Results — full data table\n"
                "  • TradingView Import — symbols for watchlist import\n"
                "  • Summary — match count and sector breakdown")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    def _export_tv(self):
        results, job_name, _ = self._current_job_info()
        if not results:
            QMessageBox.information(self, "No Data",
                "No results to export.\nRun the screener first.")
            return
        date_str = datetime.now().strftime("%Y-%m-%d")
        path, _ = QFileDialog.getSaveFileName(
            self, "Save TradingView Watchlist",
            os.path.join(BASE_DIR, f"{job_name}_{date_str}.csv"),
            "TradingView CSV (*.csv)",
        )
        if not path:
            return
        try:
            export_tradingview(results, job_name, path)
            self._log(f"TradingView watchlist saved → {path}")
            QMessageBox.information(self, "Export Successful",
                f"Watchlist saved ({len(results)} symbols):\n{path}\n\n"
                "How to import into TradingView:\n"
                "  1. Open TradingView → Watchlist panel\n"
                "  2. Click ⋮ (More options) → Import watchlist\n"
                "  3. Select this .csv file")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))


# ── Main window ────────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Trade Screener")
        self.setMinimumSize(1200, 720)
        self.resize(1480, 860)

        tabs = QTabWidget()
        tabs.setDocumentMode(True)

        self.s1 = ScreenerTab(1)
        self.s2 = ScreenerTab(2)

        self.s1.status_msg.connect(self.statusBar().showMessage)
        self.s2.status_msg.connect(self.statusBar().showMessage)

        tabs.addTab(self.s1, "📈  Screener 1 — Bullish Near Support")
        tabs.addTab(self.s2, "📉  Screener 2 — Wkly Bullish / Daily Pullback")

        self.setCentralWidget(tabs)

        sb = self.statusBar()
        sb.showMessage(
            "Ready  ·  Configure parameters on the left, then click Run Both  ·  "
            f"Settings persist in {os.path.basename(SETTINGS_FILE)}"
        )

        # Right side of status bar
        self._ver = QLabel("v2.0  ·  TradingView Data")
        self._ver.setStyleSheet("color:#4a5070; font-size:11px; margin-right:8px;")
        sb.addPermanentWidget(self._ver)


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    app = QApplication(sys.argv)
    app.setApplicationName("Trade Screener")
    app.setOrganizationName("TradeScreener")
    app.setStyle("Fusion")

    # Use the platform system font at a readable size; no custom font-family
    font = app.font()
    font.setPointSize(11)
    app.setFont(font)

    app.setStyleSheet(STYLE)

    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
